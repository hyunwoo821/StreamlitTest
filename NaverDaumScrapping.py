import requests
import base64
import time, random

import configparser
import re, shutil
import os
import pandas as pd
import math
import gc, json
import argparse
import logging
import urllib.request

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openai import OpenAI
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from datetime import datetime, timedelta
from typing import List
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

# absl 워닝 억제
os.environ["ABSL_CPP_MIN_LOG_LEVEL"] = "3"
os.environ["TF_CPP_MIN_LOG_LEVEL"] = "3"
# webdriver-manager 로그 억제
os.environ['WDM_LOG'] = '0'

# OpenAI API 키 설정
config = configparser.ConfigParser()
config.read('NewsScrappingConfig.ini')
#logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s')
logging.basicConfig(level=logging.INFO, format='%(message)s')

#openai.api_key = config['Config']['openai.api_key']
client = OpenAI(api_key=config['Config']['openai.api_key'])
####################################################################
################### Sub Task 0: 키워드 및 변수 설정 ###################
# def Variable_Setting():
#     time.sleep(1)
#     config = configparser.ConfigParser()
#     config.read('NewsScrappingConfig.ini')
#
#     #openai.api_key = config['Config']['openai.api_key']
#     client = OpenAI(api_key=config['Config']['openai.api_key'])
#     print("[Sub Task] 데이터 수집 완료")
#     return "Collected Data"

#################################################################
################### Sub Task 1: 네이버 기사 수집 ###################
def NaverCafe_Scrapping(sKeywordList, sSearchDate, sResultPath, result_file):
    time.sleep(1)

    # 폴더 존재 여부 확인 후, 없으면 생성 (검색결과 상관없이)
    output_folder = sResultPath
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # 입력받은 키워드별 반복 진행

    keyword_list = [k.strip() for k in sKeywordList.split(";") if k.strip()]
    for keyword in keyword_list:
        # 1) 크롬 드라이버 설정
        options = webdriver.ChromeOptions()
        options.add_argument("--start-maximized")  # 창 크게
        options.add_argument("--log-level=3")
        options.add_argument("--disable-background-networking")
        options.add_argument("--disable-component-update")
        driver = webdriver.Chrome(
            service=Service(ChromeDriverManager().install()),
            options=options
        )
        wait = WebDriverWait(driver, 10)

        # 2) 네이버 실행
        print("-" * 80)
        print(f"네이버 카페 검색 시작 : {keyword}")
        driver.get("https://www.naver.com")

        # 3) 검색창에 키워드 입력 후 Enter
        time.sleep(3)
        search_box = wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "input#query"))
        )
        search_box.send_keys(keyword)
        search_box.send_keys(Keys.ENTER)

        # 4) '카페' 탭 클릭
        time.sleep(5)
        cafe_tab = wait.until(EC.element_to_be_clickable((By.XPATH, "//a[@role='tab' and contains(., '카페')]")))
        cafe_tab.click()

        # 5) '최신순' 정렬 클릭
        #WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        time.sleep(3)
        latest_btn = wait.until(
            EC.element_to_be_clickable(
                (By.XPATH, "//a[contains(., '최신순')]")
            )
        )
        latest_btn.click()

        time.sleep(2)  # 정렬 반영 대기 (짧게)

        # 6) 검색결과 순차적으로 처리 (제목/링크 수집 예시)
        results = wait.until(
            EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR, "section.sc_new.sp_ncafe ul.lst_view li.bx")
            )
        )

        #print(f"검색결과 개수: {len(results)}")
        for idx, li in enumerate(results, start=1):
            # 제목 + 링크
            title_el = li.find_element(By.CSS_SELECTOR, "a.title_link")
            title = title_el.text
            url = title_el.get_attribute("href")

            # 카페명
            cafe_name_el = li.find_element(By.CSS_SELECTOR, ".user_info a.name")
            cafe_name = cafe_name_el.text

            # 날짜
            date_el = li.find_element(By.CSS_SELECTOR, ".user_info span.sub")
            date_text = date_el.text

            # 내용 요약(설명)
            try:
                desc_el = li.find_element(By.CSS_SELECTOR, ".dsc_area a.dsc_link")
                desc = desc_el.text
            except:
                desc = ""

            print(f"[{idx}]. {title}")
            print(f"네이버 카페명 : {cafe_name}")
            print(f"  작성날짜   : {date_text}")
            print(f"    링크    : {url}")
            print(f"  내용요약   : {desc[:80]}...")
            print("-" * 80)

            now = datetime.now()
            if "시간 전" in date_text:
                hours = int(re.search(r"(\d+)시간 전", date_text).group(1))
                date_obj = now - timedelta(hours=hours)
                dt_name = date_obj.strftime("%Y.%m.%d")
            elif "분 전" in date_text or "초전" in date_text:
                date_obj = now
                dt_name = date_obj.strftime("%Y.%m.%d")
            elif "일 전" in date_text:
                days = int(re.search(r"(\d+)일 전", date_text).group(1))
                date_obj = now - timedelta(days=days)
                dt_name = date_obj.strftime("%Y.%m.%d")
            elif "주 전" in date_text:
                days = int(re.search(r"(\d+)주 전", date_text).group(1))
                date_obj = now - timedelta(days=days * 7)
                dt_name = date_obj.strftime("%Y.%m.%d")
            else:
                date_text = date_text[:10]
                date_obj = datetime.strptime(date_text, "%Y.%m.%d")
                dt_name = date_text

            dt_name2 = dt_name.replace(".", "")[-6:]
            if (date_obj + timedelta(days=int(sSearchDate))).date() < now.date():
                print(f"  → {sSearchDate}일 초과, 스킵 및 종료 ({dt_name})")
                break
            else:
                print(f"  → 날짜 범위 안 계속 진행, {dt_name}")

                # ChromeDriver 실행
                driver2 = webdriver.Chrome(
                    service=Service(ChromeDriverManager().install()),
                    options=options
                )

                # 페이지 로드 최대 180초로 설정
                driver2.set_page_load_timeout(180)

                # 안전 대기 넣기 추가(25.11.28)
                driver2.get(url)
                time.sleep(3)

                # 작성날짜 가져오기(26.2.25)

                # PDF 파일 저장 경로
                title_re = sanitize_filename(title)
                pdf_path = f"네이버 카페_{keyword}_{title_re}_{dt_name2}.pdf"
                Final_path = os.path.join(output_folder, pdf_path)

                try:
                    pdf_data = driver2.execute_cdp_cmd("Page.printToPDF", {
                    })
                    time.sleep(2)
                except Exception as e:
                    print("PDF 변환 중 오류 발생:", e)
                    pdf_data = None

                # PDF 저장
                if pdf_data:
                    # PDF 저장
                    with open(Final_path, "wb") as f:
                        f.write(base64.b64decode(pdf_data["data"]))
                    print(f"PDF 저장 성공: 네이버 카페_{keyword}_{title}")

                    # 결과 입력
                    dt_name = dt_name.replace(".", "-")
                    Excel_EnterResult(result_file, "네이버 카페", keyword, title_re, dt_name, url)
                else:
                    print(f"PDF 저장 실패: 네이버 카페_{keyword}_{title}")

                driver2.quit()
                time.sleep(2)

        # 다음 키워드 검색을 위한 드라이버 재실행 전 대기
        driver.quit()
        time.sleep(5)
        print(f"네이버 카페 검색 종료 : {keyword}")

    print(f"네이버 카페 전체 종료")
    print("-" * 80)
################################################################
################### Sub Task 2: 다음 기사 수집 ###################
def NaverBlog_Scrapping(sKeywordList, sSearchDate, sResultPath, result_file):
    #네이버 API 기본값 세팅
    client_id = "CJnSJvwL77q34PRPxoQX"
    client_secret = "wL870ijima"

    # 폴더 존재 여부 확인 후, 없으면 생성 (검색결과 상관없이)
    output_folder = sResultPath
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # 입력받은 키워드별 반복 진행
    keyword_list = [k.strip() for k in sKeywordList.split(";") if k.strip()]
    for keyword in keyword_list:
        bExitFlag = False
        encText = urllib.parse.quote(keyword)


        #10개페이지 100개 게시물, 최대 1000개 게시물 조회
        for pageCnt in range(1, 1001, 100):
            url = ("https://openapi.naver.com/v1/search/blog?query="  # query: 검색어, display: 표시할 검색 결과, start : 검색 시작위치
                   + encText + "&display=100&start=" + str(pageCnt) + "&sort=date")  # sort: sim(정확도순 내림차순), date(날짜순 내림차순)
            request = urllib.request.Request(url)
            request.add_header("X-Naver-Client-Id", client_id)
            request.add_header("X-Naver-Client-Secret", client_secret)
            response = urllib.request.urlopen(request)
            rescode = response.getcode()
            time.sleep(1)

            print("-" * 80)
            print(f"네이버 블로그 검색 시작 : {keyword}")
            if (rescode == 200):
                response_body = response.read()
            else:
                print("Error Code:" + rescode + ", 해당 키워드 종료")
                break

            data = response_body.decode('utf-8')
            data = json.loads(data)

            results = []

            for idx, item in enumerate(data.get("items", []), start=1):
                title = item.get("title", "").strip()               # 게시물 제목
                blog_name = item.get("bloggername", "").strip()     # 블로그명
                url = item.get("link", "").strip()                  # URL
                desc = item.get("description", "").strip()          # 내용
                postdate_raw = item.get("postdate", "").strip()     # 작성날짜 (YYYYMMDD)

                # 날짜 포맷 변환 (선택)
                if postdate_raw:
                    dt_name = datetime.strptime(postdate_raw, "%Y%m%d").strftime("%Y-%m-%d")

                print(f"[{idx}]. {title}")
                print(f" 네이버 블로그명 : {blog_name}")
                print(f"    작성날짜   : {dt_name}")
                print(f"     링크     : {url}")
                print(f"    내용요약   : {desc[:80]}...")
                print("-" * 80)

                now = datetime.today().date()
                tmp_date = datetime.strptime(postdate_raw, "%Y%m%d").date()
                diff_days = (now - tmp_date).days
                if int(sSearchDate) < diff_days:
                    print(f"  → {sSearchDate}일 초과, 스킵 및 종료")
                    bExitFlag = True
                    break
                else:
                    print(f"  → 날짜 범위 안 계속 진행, {diff_days}")

                    # 크롬 드라이버 설정
                    options = webdriver.ChromeOptions()
                    options.add_argument("--start-maximized")  # 창 크게
                    options.add_argument("--log-level=3")
                    options.add_argument("--disable-background-networking")
                    options.add_argument("--disable-component-update")
                    driver = webdriver.Chrome(
                        service=Service(ChromeDriverManager().install()),
                        options=options
                    )

                    # 페이지 로드 최대 180초로 설정
                    driver.set_page_load_timeout(180)

                    # 화면 전체 인쇄를 위해 모바일 버전으로 열기
                    url = url.replace("https://", "https://m.")
                    driver.get(url)

                    # 인쇄하기 위한 화면 조절 셋팅 시작
                    # 본문 컨테이너 찾기
                    article = driver.find_element(By.CSS_SELECTOR, ".se-main-container")

                    # 본문 높이 계산
                    article_height = driver.execute_script(
                        "return arguments[0].scrollHeight;", article
                    )

                    # 본문 끝까지만 스크롤
                    driver.execute_script(
                        "window.scrollTo(0, arguments[0]);", article_height
                    )

                    # 이미지 로딩 대기
                    time.sleep(2)

                    driver.execute_script("""
                    document.querySelectorAll('img').forEach(img => {
                      img.removeAttribute('loading');
                    });
                    """)

                    # PDF 파일 저장 경로
                    dt_name2 = dt_name.replace(".", "")[-6:]
                    title_re = sanitize_filename(title)
                    pdf_path = f"네이버 블로그_{keyword}_{title_re}_{dt_name2}.pdf"
                    Final_path = os.path.join(output_folder, pdf_path)

                    try:
                        pdf_data = driver.execute_cdp_cmd("Page.printToPDF", {
                            "printBackground": True,
                            "preferCSSPageSize": True,  # 사이트 CSS @page 있으면 그거 우선
                            "scale": 1.0,
                            "marginTop": 0.3,
                            "marginBottom": 0.3,
                            "marginLeft": 0.3,
                            "marginRight": 0.3,
                            "paperWidth": 8.27,  # A4 (inch)
                            "paperHeight": 11.69,  # A4 (inch)
                            "landscape": False
                        })
                        time.sleep(2)
                    except Exception as e:
                        print("PDF 변환 중 오류 발생:", e)
                        pdf_data = None

                    # PDF 저장
                    if pdf_data:
                        # PDF 저장
                        with open(Final_path, "wb") as f:
                            f.write(base64.b64decode(pdf_data["data"]))
                        print(f"PDF 저장 성공: 네이버 블로그_{keyword}_{title}")

                        # 결과 입력
                        Excel_EnterResult(result_file, "네이버 블로그", keyword, title, dt_name, url)
                    else:
                        print(f"PDF 저장 실패: 네이버 블로그_{keyword}_{title}")

                    driver.quit()
                    time.sleep(2)

            if bExitFlag == True:
                print(f"네이버 블로그 검색 종료 : {keyword}")
                break
            else:
                print(f"네이버 블로그 키워드{keyword} {pageCnt} +100개 완료, 다음 페이지 진행..")

    print(f"네이버 블로그 전체 종료")
    print("-" * 80)
################################################################
################### Sub Task 3: 다음 통합 검색 ###################
def DaumSearch_Scrapping(sKeywordList, sSearchDate, sResultPath, result_file):

    def format_date(dt_str):
        if not dt_str:
            return ""
        return datetime.fromisoformat(dt_str).strftime("%Y-%m-%d")

    def clean_html(text):
        if not text:
            return ""
        return re.sub(r"<.*?>", "", text)

    DOWNLOAD_HINTS = [
        "download",
        "filedown",
        "fileDownload",
        "attach",
        "att_no=",
        "atchFileId",
        "filename=",
        "fileid",
        "f_key",
        "mode=download",
        ".pdf",
        ".hwp",
        ".zip",
        ".xls",
        ".ppt",
        "namu.wiki"
    ]

    def is_suspected_download_url(url):
        url = url.lower()
        return any(hint in url for hint in DOWNLOAD_HINTS)

    # 폴더 존재 여부 확인 후, 없으면 생성 (검색결과 상관없이)
    output_folder = sResultPath
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # 입력받은 키워드별 반복 진행
    keyword_list = [k.strip() for k in sKeywordList.split(";") if k.strip()]
    for keyword in keyword_list:
        PassDate = False
        # 다음(카카오) API 기본값 세팅
        REST_API_KEY = "d4c1113ae8ccc31bddbe5cdf4884237b"
        query = keyword
        url = "https://dapi.kakao.com/v2/search/web"
        headers = {
            "Authorization": f"KakaoAK {REST_API_KEY}"
        }
        print("-" * 80)
        print(f"다음 통합 검색 시작 : {keyword}")

        for Cnt in range(1, 20):
            params = {
                "query": query,     # 검색어
                "sort": "recency",  # recency는 최신순, accuracy는 정확도순
                "page": Cnt,        # 결과 페이지 번호
                "size": 50          # 한 페이지에 보여질 문서 수
            }

            response = requests.get(url, headers=headers, params=params)

            if response.status_code == 200:
                result = response.json()
                documents = result.get("documents", [])

                for idx, doc in enumerate(documents, start=1):
                    url = doc.get("url", "")                                # URL
                    title = clean_html(doc.get("title", ""))                # 제목
                    date_obj = format_date(doc.get("datetime", ""))         # 날짜
                    dt_name = date_obj.replace("-", "")[-6:]    # 날짜 (yyMMdd)

                    # 입력받은 날짜 초과 작성일자 기사 스킵
                    post_date = datetime.strptime(date_obj, "%Y-%m-%d").date()
                    today = datetime.today().date()
                    standard = today - timedelta(days=int(sSearchDate))
                    if post_date < standard:
                        print(f"작성일자가 {sSearchDate}일 초과, 스킵 및 종료")
                        PassDate = True
                        break
                    else:
                        if is_suspected_download_url(url):
                            print(f"{idx}. 비정상URL(다운로드 파일) 로 SKIP처리 {url}")
                        else:
                            print(f"[{idx}]. {title}")
                            print(f"   링크 : {url}")
                            print(f" 작성날짜  : {date_obj}")
                            print("-" * 80)

                            # 크롬 드라이버 설정
                            options = webdriver.ChromeOptions()
                            options.add_argument("--start-maximized")  # 창 크게
                            options.add_argument("--log-level=3")
                            options.add_argument("--disable-background-networking")
                            options.add_argument("--disable-component-update")
                            driver2 = webdriver.Chrome(
                                service=Service(ChromeDriverManager().install()),
                                options=options
                            )

                            # 페이지 로드 최대 180초로 설정
                            driver2.set_page_load_timeout(180)

                            # 안전 대기 넣기 추가
                            driver2.get(url)
                            time.sleep(3)

                            # PDF 파일 저장 경로
                            title_re = sanitize_filename(title)
                            pdf_path = f"다음 통합 검색_{keyword}_{title_re}_{dt_name}.pdf"
                            Final_path = os.path.join(output_folder, pdf_path)

                            try:
                                pdf_data = driver2.execute_cdp_cmd("Page.printToPDF", {
                                })
                                time.sleep(2)
                            except Exception as e:
                                print("PDF 변환 중 오류 발생:", e)
                                pdf_data = None

                            # PDF 저장
                            if pdf_data:
                                # PDF 저장
                                with open(Final_path, "wb") as f:
                                    f.write(base64.b64decode(pdf_data["data"]))
                                print(f"PDF 저장 성공: 다음 통합 검색_{keyword}_{title}")

                                # 결과 입력
                                Excel_EnterResult(result_file, "다음 통합 검색", keyword, title_re, date_obj, url)
                            else:
                                print(f"PDF 저장 실패: 다음 통합 검색_{keyword}_{title}")

                            driver2.quit()
                            time.sleep(2)

                if PassDate == True:
                    print(f"다음 통합 검색 종료 : {keyword}")
                    break
            else:
                print("Error로 인한 종료: ", response.status_code, response.text)
                break

            if PassDate == True:
                break

    print(f'다음 통합 검색 전체 종료')
    print("-" * 80)
###################################################################

def fetch_article_content(url):
    # 주어진 URL에서 기사 내용을 가져옵니다.
    try:
        #원본 기사 내용 가져오기
        response = requests.get(url)
        response.raise_for_status()  # HTTP 상태 코드 확인
        soup = BeautifulSoup(response.text, "html.parser")

        # 기사 내용 추출 (웹사이트 구조에 따라 태그 변경 필요)
        paragraphs = soup.find_all('p')
        content = "\n".join([para.get_text() for para in paragraphs])
        return content
    except Exception as e:
        return f"기사 내용을 가져오는 중 오류 발생: {str(e)}"


## Common Task 2: Page PDF Download
# def PDF_Download():
#     time.sleep(1)
#     print(f"[Sub Task] '{result}' 저장 완료")
#     return f"Saved {result}"

# Common Task 3: Enter Result Information
def Excel_EnterResult(result_file, sSearchSite, sKeyWord, sSearchTitle, date_text, sSearchURL):
    time.sleep(1)

    # 기존 결과 파일 로드
    df_result = pd.read_excel(result_file)

    # 연번 계산
    next_no = len(df_result) + 1

    # 추가할 행
    new_row = {
        '연번': next_no,
        '사이트명': sSearchSite,
        '키워드': sKeyWord,
        '제목': sSearchTitle,
        '작성날짜': date_text,
        '링크': sSearchURL
    }

    # 행 추가
    df_result.loc[len(df_result)] = new_row

    # 저장
    df_result.to_excel(result_file, index=False)

    # ===== 엑셀 서식 처리 (openpyxl) =====
    wb = load_workbook(result_file)
    ws = wb.active


    # 열 너비 자동조정 (개선버전)
    for column_cells in ws.columns:
        column_letter = column_cells[0].column_letter
        max_length = 0

        for cell in column_cells:
            if cell.value:
                cell_value = str(cell.value)

                # 한글은 1.3배 가중치 적용
                adjusted_length = 0
                for ch in cell_value:
                    if ord(ch) > 127:   # 한글/유니코드 문자
                        adjusted_length += 1.3
                    else:
                        adjusted_length += 1

                max_length = max(max_length, adjusted_length)

        # 최소 10, 최대 60 제한
        final_width = min(max(math.ceil(max_length) + 4, 10), 60)

        ws.column_dimensions[column_letter].width = final_width

    # 🔹 방금 추가된 행 가운데 정렬
    target_row = next_no + 1  # 1행은 헤더

    for cell in ws[target_row]:
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # # A~E 열 자동 너비 조정
    # for col in ['A', 'B', 'C', 'D', 'E']:
    #     max_length = 0
    #     for cell in ws[col]:
    #         if cell.value:
    #             max_length = max(max_length, len(str(cell.value)))
    #     ws.column_dimensions[col].width = max_length + 2
    #
    # # 2️방금 추가된 행(2행부터 누적이므로 실제 행 번호는 next_no + 1)
    # target_row = next_no + 1  # 1행은 헤더
    #
    # for col in ['A', 'B', 'C', 'D', 'E']:
    #     ws[f"{col}{target_row}"].alignment = Alignment(horizontal="center", vertical="center")

    wb.save(result_file)

# Common Task 4 : Change special characters
def sanitize_filename(title):
    """파일명으로 사용할 수 없는 문자를 공백(Sapce)으로 변환"""
    return re.sub(r'[\/:*?"<>|]', ' ', title)

def dedup_news_pdfs(test_root, keep="newest", dry_run=True, safe_trash=True):
  """
  test_root: Test 폴더 경로
  keep:     "newest"(기본, yyMMdd 가장 최신 보존) | "first"(첫 파일 보존)
  dry_run:  True면 실제 삭제/이동하지 않고 로그만 출력
  safe_trash: True면 삭제 대신 회사폴더 내 __DUP_TRASH로 이동
  """

  root = Path(test_root)
  if not root.exists():
    raise FileNotFoundError(f"경로 없음: {root}")

  def norm_title(t):
    # 한글/영문만 남기고 소문자화
    return "".join(re.findall(r"[가-힣A-Za-z]+", t)).lower()

  def parse(stem):
    # 파일명: 건설사명_yyMMdd_기사제목  → (company, yymmdd, title_raw)
    parts = stem.split("_", 2)
    if len(parts) >= 3 and re.fullmatch(r"\d{6}", parts[1] or ""):
      return parts[0], parts[1], parts[2]
    # 패턴 불일치 시 제목만 사용
    return None, None, stem

  def sortkey(yymmdd):
    # 날짜 없으면 가장 오래된 취급
    return (yymmdd is None, yymmdd or "000000")

  total_scanned = total_dups = total_deleted = 0
  print(f"=== 시작: {root} ===")

  # Test 폴더 바로 아래 '회사 폴더'만 처리
  for company_dir in [p for p in root.iterdir() if p.is_dir()]:
    # 확장자 대소문자 모두 처리
    pdfs = [p for p in company_dir.iterdir() if p.is_file() and p.suffix.lower() == ".pdf"]
    if not pdfs:
      continue

    # 제목(정규화)별 그룹핑: {title_key: [(yymmdd, Path), ...]}
    groups = {}
    for p in pdfs:
      _, yymmdd, title_raw = parse(p.stem)
      title_k = norm_title(title_raw)
      groups.setdefault(title_k, []).append((yymmdd, p))

    scanned = sum(len(v) for v in groups.values())
    dups = deleted = 0

    for _, files in groups.items():
      if len(files) == 1:
        continue

      # 보존 파일 결정
      if keep == "newest":
        keep_item = max(files, key=lambda x: sortkey(x[0]))
      else:
        keep_item = files[0]

      to_remove = [p for (d, p) in files if p != keep_item[1]]
      dups += len(to_remove)

      # 삭제(또는 휴지통 이동)
      for p in to_remove:
        if safe_trash:
          trash = company_dir / "__DUP_TRASH"
          target = trash / p.name
          if not dry_run:
            trash.mkdir(exist_ok=True)
            # 이름 충돌 방지
            if target.exists():
              i = 1
              while True:
                cand = trash / f"{p.stem}__dup{i}{p.suffix}"
                if not cand.exists():
                  target = cand
                  break
                i += 1
            shutil.move(str(p), str(target))
          print(f"[TRASH] {p.name}")
        else:
          if not dry_run:
            p.unlink(missing_ok=True)  # Py3.9 지원
          print(f"[DEL]   {p.name}")
        deleted += 1

    total_scanned += scanned
    total_dups += dups
    total_deleted += deleted
    print(f"[{company_dir.name}] 스캔:{scanned}  중복:{dups}  삭제:{deleted}")

  print(f"\n=== 전체 ===  스캔:{total_scanned}  중복:{total_dups}  삭제:{total_deleted}")
  if dry_run:
    print("※ DRY_RUN=True: 실제 삭제/이동 없음")

# Main Task: 전체 프로세스 실행
def main_task():
    print("[Main Task] 시작")

    parser = argparse.ArgumentParser(
        description="자격증 대여 검색 자동화"
    )
    parser.add_argument(
        "sSearchSites",
        help="스크래핑할 사이트 목록(‘;’으로 구분 ex: 네이버 카페;네이버 블로그)"
    )
    parser.add_argument(
        "sKeyword",
        help="사이트에서 검색할 키워드"
    )
    parser.add_argument(
        "sSearchDate",
        help="검색허용할 날짜 기간"
    )
    parser.add_argument(
        "result_path",
        help="결과를 저장할 폴더 경로"
    )
    args = parser.parse_args()
    sSearchSites = args.sSearchSites
    site_list = [s.strip() for s in sSearchSites.split(';') if s.strip()]

    sKeyword = args.sKeyword

    print(f"스크래핑 대상 사이트: {sSearchSites}")
    print(f"검색할 키워드: {sKeyword}")
    print(f"검색 허용 기간: {args.sSearchDate}일")
    print(f"결과 폴더: {args.result_path}")

    # 초기변수 설정
    #Variable_Setting()  # Sub 0 Task

    # 결과 파일 생성 및 헤더 추가
    result_file = os.path.join(args.result_path, 'Result.xlsx')
    df_result = pd.DataFrame(columns=['연번', '사이트명', '키워드', '제목', '작성날짜', '링크'])
    df_result.to_excel(result_file, index=False)

    #pyautogui.alert('강제 멈추기.')
    # 입력한 검색 사이트에 네이버 카페가 있는 경우
    if any('네이버 카페' in str(value) for value in site_list):
        NaverCafe_Scrapping(sKeyword, args.sSearchDate, args.result_path, result_file)    # Sub 1 Task
    # 입력한 검색 사이트에 네이버 블로그가 있는 경우
    if any('네이버 블로그' in str(value) for value in site_list):
        NaverBlog_Scrapping(sKeyword, args.sSearchDate, args.result_path, result_file)     # Sub 2 Task
    # 입력한 검색 사이트에 다음 통합 검색이 있는 경우
    if any('다음 통합 검색' in str(value) for value in site_list):
        DaumSearch_Scrapping(sKeyword, args.sSearchDate, args.result_path, result_file)  # Sub 3 Task

    final_result = 2

    print("[뉴스 검색 스크래핑] 전체 종료")

    #dedup_news_pdfs(args.result_path, keep="newest", dry_run=False, safe_trash=True)
    #dedup_news_pdfs(args.result_path, keep="newest", dry_run=True, safe_trash=False)
    #print("중복기사 삭제 완료")
    return final_result

# 실행
if __name__ == "__main__":
    # sSearchSites = input('사이트 입력: ')
    # file_path = input('엑셀 파일 경로를 입력하세요: ')
    # result_path = input('결과 저장 경로를 입력하세요:')
    # print("검새할 사이트리스트:", sSearchSites)
    # print("엑셀파일 경로:", file_path)
    # print("결과 저장 경로:", result_path)
    #main_task(sSearchSites, file_path, result_path)
    # print("종료되었습니다.")

    main_task()
